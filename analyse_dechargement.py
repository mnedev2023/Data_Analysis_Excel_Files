#!/usr/bin/env python3
"""
Analyse et recalcul des données de déchargement de navires à partir d'un
fichier Excel.

Nouveautés :
-----------
* Les colonnes numériques suivantes sont **formatées dans Excel** avec
  deux décimales et séparateur de milliers (« #,##0.00 ») :

    - Poids eau Calculé (kg)
    - Poids net Calculé (kg)
    - Poids net Recalculé (kg)

* Ajustement automatique de la largeur des colonnes (inchangé).
* Export du résultat vers « Data_Analysis/YYYY‑MM‑DD/…_resultats.xlsx ».

Usage :
    python analyse_dechargement.py                # boîte de sélection
    python analyse_dechargement.py --input chemin/fichier.xlsx
"""

from __future__ import annotations

from pathlib import Path
from datetime import date
import os
import sys
import subprocess
import argparse
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Noms de colonnes attendues dans le fichier source
# ---------------------------------------------------------------------------
COL_DEBUT_PESEE: str = "Date heure 1ère pesée"
COL_FIN_PESEE: str = "Date heure 2ème pesée"
COL_DEBUT_DECH: str = "Date début déchargement"
COL_FIN_DECH: str = "Date fin déchargement"

COL_DUREE_OPERATION: str = "Durée opération"
COL_TEMPS_TR: str = "Temps traitement"

COL_VOL_INIT: str = "Volume initial"
COL_VOL_FINAL: str = "Volume final"

COL_POIDS_ENTREE: str = "Poids entrée (kg)"
COL_POIDS_SORTIE: str = "Poids sortie (kg)"
COL_POIDS_EAU: str = "Poids eau (kg)"  # mesuré dans le fichier d'origine

# ---------------------------------------------------------------------------
# Noms des nouvelles colonnes générées par ce script
# ---------------------------------------------------------------------------
COL_VOL_CHARGE_CALCULE: str = "Volume chargé (m³)"
COL_POIDS_EAU_CALCULE: str = "Poids eau Calculé (kg)"
COL_POIDS_NET_CALCULE: str = "Poids net Calculé (kg)"  # basé sur poids eau mesuré
COL_POIDS_NET_RECALCULE: str = "Poids net Recalculé (kg)"  # basé sur poids eau recalculé

NUMERIC_COLUMNS_TO_FORMAT = [
    COL_POIDS_EAU_CALCULE,
    COL_POIDS_NET_CALCULE,
    COL_POIDS_NET_RECALCULE,
]

# Modèle minimal pour créer un fichier d'exemple
COLONNES_MODELE: list[str] = [
    COL_DEBUT_PESEE,
    COL_FIN_PESEE,
    COL_DEBUT_DECH,
    COL_FIN_DECH,
    COL_VOL_INIT,
    COL_VOL_FINAL,
    COL_POIDS_ENTREE,
    COL_POIDS_SORTIE,
    COL_POIDS_EAU,  # colonne facultative mais recommandée si mesure disponible
]


# ---------------------------------------------------------------------------
# Fonctions utilitaires
# ---------------------------------------------------------------------------
def appliquer_format_numerique(wb_path: Path, header_row: int = 1) -> None:
    """Applique un format numérique (#,##0.00) aux colonnes listées."""
    wb = load_workbook(wb_path)
    ws = wb.active

    # Associer nom de colonne → index Excel
    col_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[header_row])}
    num_format = "#,##0.00"  # 2 décimales + séparateur milliers

    for col_name in NUMERIC_COLUMNS_TO_FORMAT:
        idx = col_index.get(col_name)
        if idx is None:
            continue  # colonne absente
        letter = get_column_letter(idx)
        for row in range(header_row + 1, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = num_format
    wb.save(wb_path)

def ajuster_largeur_colonnes(fichier_excel: Path) -> None:
    """Ajuste la largeur des colonnes en fonction du contenu."""
    wb = load_workbook(fichier_excel)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                valeur = str(cell.value) if cell.value is not None else ""
            except Exception:
                valeur = ""
            max_length = max(max_length, len(valeur))
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(fichier_excel)


def ouvrir_fichier(chemin_fichier: Path) -> None:
    """Ouvre le fichier dans l'application associée (Windows, macOS, Linux)."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(chemin_fichier)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.call(["open", chemin_fichier])
        else:
            subprocess.call(["xdg-open", chemin_fichier])
    except Exception as exc:
        print(f"⚠️ Impossible d'ouvrir le fichier automatiquement : {exc}")


# ---------------------------------------------------------------------------
# Cœur de l'analyse
# ---------------------------------------------------------------------------

def analyser_dechargement(fichier_excel: Path, dossier_sortie: Path) -> Path:
    """Lit *fichier_excel*, calcule les nouvelles métriques et sauvegarde le résultat.

    Parameters
    ----------
    fichier_excel : Path
        Le chemin du fichier source.
    dossier_sortie : Path
        Dossier dans lequel enregistrer le fichier résultat.

    Returns
    -------
    Path
        Chemin du fichier résultat créé.
    """
    # ---------------------------------------------------------------------
    # 1) Lecture du fichier d'origine
    # ---------------------------------------------------------------------
    parse_cols = [COL_DEBUT_PESEE, COL_FIN_PESEE, COL_DEBUT_DECH, COL_FIN_DECH]
    df = pd.read_excel(
        fichier_excel,
        parse_dates=parse_cols,
        dtype="object",  # on laisse pandas déterminer les types ensuite
    )

    # Nettoyage des en‑têtes (espaces superflus)
    df.columns = df.columns.str.strip()

    print("📊 Aperçu du tableau Excel importé :")
    print(df.head())

    # ---------------------------------------------------------------------
    # 2) Calculs
    # ---------------------------------------------------------------------
    df[COL_VOL_CHARGE_CALCULE] = round(df[COL_VOL_FINAL] - df[COL_VOL_INIT], 2)
    df[COL_POIDS_EAU_CALCULE] = round(df[COL_VOL_CHARGE_CALCULE] * (1 + 6.6 / 100), 2)

    df[COL_DUREE_OPERATION] = df[COL_FIN_DECH] - df[COL_DEBUT_DECH]
    df[COL_TEMPS_TR] = df[COL_FIN_PESEE] - df[COL_DEBUT_PESEE]

    # Poids net calculé (utilise la mesure de l'eau si disponible)
    df[COL_POIDS_NET_CALCULE] = round(
        (df[COL_POIDS_SORTIE] - df[COL_POIDS_ENTREE] - df[COL_POIDS_EAU]) * (1 - 7 / 100), 2
    )


    # Poids net recalculé (utilise le poids eau recalculé)
    df[COL_POIDS_NET_RECALCULE] = round(
        (df[COL_POIDS_SORTIE] - df[COL_POIDS_ENTREE] - df[COL_POIDS_EAU_CALCULE]) * (1 - 7 / 100), 2
    )
    print("📊 Tableau final avec colonnes calculées :")
    print(df.head(10))

    # ---------------------------------------------------------------------
    # 3) Export du résultat ( Sauvegarde)
    # ---------------------------------------------------------------------
    dossier_sortie.mkdir(parents=True, exist_ok=True)
    fichier_resultat = dossier_sortie / f"{fichier_excel.stem}_resultats.xlsx"
    df.to_excel(fichier_resultat, index=False)
    appliquer_format_numerique(fichier_resultat)
    ajuster_largeur_colonnes(fichier_resultat)

    print(f"✔︎ Résultat enregistré : {fichier_resultat.relative_to(Path.cwd()) if fichier_resultat.is_relative_to(Path.cwd()) else fichier_resultat}")
    return fichier_resultat


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def cli_selectionner_fichier() -> Path | None:
    """Ouvre une fenêtre de sélection de fichier et renvoie le chemin choisi."""
    root = tk.Tk()
    root.withdraw()
    dossier_defaut = Path.cwd() / "Excel"
    dossier_defaut.mkdir(parents=True, exist_ok=True)
    chemin = filedialog.askopenfilename(
        title="Sélectionner un fichier Excel",
        initialdir=dossier_defaut,
        filetypes=[("Fichiers Excel", "*.xlsx *.xls")],
    )
    root.destroy()
    return Path(chemin) if chemin else None


def creer_fichier_modele(dossier_excel: Path) -> None:
    """Crée un fichier modèle vierge dans *dossier_excel* si nécessaire."""
    fichier_modele = dossier_excel / "modele_import.xlsx"
    if fichier_modele.exists():
        return
    pd.DataFrame(columns=COLONNES_MODELE).to_excel(fichier_modele, index=False)
    print(f"📄 Fichier modèle créé : {fichier_modele.relative_to(Path.cwd()) if fichier_modele.is_relative_to(Path.cwd()) else fichier_modele}")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Analyse de déchargement de navires")
    parser.add_argument("--input", "-i", type=Path, help="Chemin du fichier Excel à traiter")
    args = parser.parse_args(argv)

    dossier_excel = Path.cwd() / "Excel"
    dossier_excel.mkdir(parents=True, exist_ok=True)

    if args.input is None:
        # Aucun fichier passé en argument → boîte de dialogue
        if not any(dossier_excel.glob("*.xlsx")):
            print("📁 Aucun fichier trouvé dans le dossier ‘Excel/’.")
            creer_fichier_modele(dossier_excel)
            print("📝 Place tes fichiers dans ce dossier ou passe le chemin en argument, puis relance le script.")
            return
        fichier_source = cli_selectionner_fichier()
        if fichier_source is None:
            print("❌ Aucun fichier sélectionné. Fin du programme.")
            return
    else:
        fichier_source = args.input
        if not fichier_source.exists():
            print(f"❌ Fichier introuvable : {fichier_source}")
            return

    # Dossier de sortie daté
    dossier_sortie = Path.cwd() / "Data_Analysis" / date.today().isoformat()

    try:
        fichier_resultat = analyser_dechargement(fichier_source, dossier_sortie)
        ouvrir_fichier(fichier_resultat)
    except Exception as exc:
        print(f"⚠️ Erreur lors du traitement : {exc}")


if __name__ == "__main__":
    main()
